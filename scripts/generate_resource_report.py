import json
import os

from openpyxl import Workbook
from openpyxl.styles import PatternFill, Font

# =========================================================
# CONFIG
# =========================================================
BUILD_ID = str(os.environ.get("BUILD_BUILDID", "")).strip()

INPUT_FILE = "output/resources.json"

OUTPUT_DIR = "output"

OUTPUT_FILE = f"{OUTPUT_DIR}/Resource_List_Hexa_Training.xlsx"

# =========================================================
# LOAD RESOURCE JSON
# =========================================================
if not os.path.exists(INPUT_FILE):
    raise FileNotFoundError(
        f"{INPUT_FILE} not found."
    )

with open(INPUT_FILE, "r", encoding="utf-8") as f:
    resources = json.load(f)

# =========================================================
# CREATE WORKBOOK
# =========================================================
wb = Workbook()

ws = wb.active

ws.title = "Azure Resources"

# =========================================================
# HEADERS
# =========================================================
headers = [
    "S No.",
    "Resource Group Name",
    "Resource Name",
    "Resource Type",
    "ENV",
    "Status"
]

ws.append(headers)

# =========================================================
# STYLES
# =========================================================
header_font = Font(
    bold=True
)

header_fill = PatternFill(
    start_color="1F4E78",
    end_color="1F4E78",
    fill_type="solid"
)

header_text = Font(
    bold=True,
    color="FFFFFF"
)

# NEW RESOURCE STYLE
new_fill = PatternFill(
    start_color="008000",
    end_color="008000",
    fill_type="solid"
)

new_font = Font(
    bold=True,
    color="90EE90"  # Light Green
)

# =========================================================
# APPLY HEADER STYLE
# =========================================================
for col in range(1, len(headers) + 1):

    cell = ws.cell(row=1, column=col)

    cell.font = header_text

    cell.fill = header_fill

# =========================================================
# SORT RESOURCES
# =========================================================
resources = sorted(
    resources,
    key=lambda x: (
        x.get("resourceGroup", ""),
        x.get("name", "")
    )
)

# =========================================================
# PROCESS RESOURCES
# =========================================================
for index, res in enumerate(resources, start=1):

    name = res.get("name", "")

    rg = res.get("resourceGroup", "")

    rtype = res.get("type", "")

    env = res.get("env", "")

    resource_build_id = str(
        res.get("buildId", "")
    ).strip()

    # =====================================================
    # DETERMINE STATUS
    # =====================================================
    if (
        BUILD_ID != ""
        and resource_build_id == BUILD_ID
    ):
        status = "New"
    else:
        status = "Existing"

    # =====================================================
    # ADD ROW
    # =====================================================
    row = [
        index,
        rg,
        name,
        rtype,
        env,
        status
    ]

    ws.append(row)

    current_row = ws.max_row

    # =====================================================
    # HIGHLIGHT NEW RESOURCES
    # =====================================================
    if status == "New":

        for col in range(1, len(headers) + 1):

            cell = ws.cell(
                row=current_row,
                column=col
            )

            cell.fill = new_fill

            cell.font = new_font

# =========================================================
# AUTO COLUMN WIDTH
# =========================================================
for column_cells in ws.columns:

    max_length = max(
        len(str(cell.value or ""))
        for cell in column_cells
    )

    adjusted_width = max_length + 5

    ws.column_dimensions[
        column_cells[0].column_letter
    ].width = adjusted_width

# =========================================================
# FREEZE HEADER
# =========================================================
ws.freeze_panes = "A2"

# =========================================================
# CREATE OUTPUT DIRECTORY
# =========================================================
os.makedirs(
    OUTPUT_DIR,
    exist_ok=True
)

# =========================================================
# SAVE FILE
# =========================================================
wb.save(OUTPUT_FILE)

print(f"Excel report generated: {OUTPUT_FILE}")